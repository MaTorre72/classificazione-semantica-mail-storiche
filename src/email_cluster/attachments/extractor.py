from __future__ import annotations

import csv
import io
from pathlib import Path


def extract_attachment_text(
    payload: bytes, filename: str | None, mime_type: str | None, max_size_mb: int = 20,
) -> tuple[str | None, str, str | None]:
    if len(payload) > max_size_mb * 1024 * 1024:
        return None, "too_large", None
    suffix = Path(filename or "").suffix.lower()
    try:
        if suffix in {".txt", ".csv"} or (mime_type or "").startswith("text/"):
            text = payload.decode("utf-8", errors="replace")
            if suffix == ".csv":
                rows = list(csv.reader(io.StringIO(text)))[:100]
                text = "\n".join(" | ".join(row[:20]) for row in rows)
            return text[:100_000], "extracted", None
        if suffix == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(payload))
            text = "\n".join((page.extract_text() or "") for page in reader.pages[:50])
            return text[:100_000], "extracted" if text.strip() else "empty", None
        if suffix == ".docx":
            from docx import Document

            document = Document(io.BytesIO(payload))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return text[:100_000], "extracted" if text.strip() else "empty", None
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in workbook.worksheets[:10]:
                lines.append(f"Foglio: {sheet.title}")
                for row in sheet.iter_rows(max_row=100, max_col=20, values_only=True):
                    values = [str(value) for value in row if value not in (None, "")]
                    if values:
                        lines.append(" | ".join(values))
            text = "\n".join(lines)
            return text[:100_000], "extracted" if text.strip() else "empty", None
    except ImportError as exc:
        return None, "dependency_missing", str(exc)
    except Exception as exc:  # noqa: BLE001 - malformed third-party documents stay isolated
        return None, "error", str(exc)
    return None, "unsupported", None
